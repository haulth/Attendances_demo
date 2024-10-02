
from openpyxl import load_workbook, Workbook
import pandas as pd
import json, os, csv, requests, shutil, re, io
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from django.core.mail import send_mail, EmailMessage
from django.shortcuts import (
    HttpResponse,
    get_object_or_404,
    redirect,
    render,
)
from django.templatetags.static import static
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from .forms import *
from .models import *
import qrcode
from io import BytesIO
from django.conf import settings
from datetime import datetime
from django.utils import timezone
import pytz


def admin_home(request):
    total_staff = Staff.objects.count()
    total_students = Student.objects.count()
    total_course = Course.objects.count()
    total_subject = Subject.objects.count()
    total_teaching = TeachingSchedule.objects.count()
    subjects = Subject.objects.all()
    courses = Course.objects.all()
    

    # Lấy ngày hôm nay
    vn_tz = pytz.timezone("Asia/Ho_Chi_Minh")
    selected_date = timezone.now().astimezone(vn_tz).date()

    print(f"ngày được chọn: {selected_date}")

    # Filter logic based on user inputs
    selected_course = request.GET.get('session', None)  # Selected school (course)
    selected_subject = request.GET.get('subject', None)  # Selected class (subject)
    start_date = request.GET.get('start_date', None)
    end_date = request.GET.get('end_date', None)

    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        # Default to today
        vn_tz = pytz.timezone("Asia/Ho_Chi_Minh")
        today = timezone.now().astimezone(vn_tz).date()
        start_date = today
        end_date = today

    filtered_attendance = Attendance.objects.filter(date__range=[start_date, end_date])

    if selected_course:
        filtered_attendance = filtered_attendance.filter(subject__course_id=selected_course)
    if selected_subject:
        filtered_attendance = filtered_attendance.filter(subject_id=selected_subject)
        
    
    # Dữ liệu danh sách tên các lớp và số học sinh điểm danh hôm nay
    subject_list = []
    student_present_today_list = []
    student_absent_today_list = []  # Danh sách số học sinh vắng mặt theo lớp
    course_student_counts = {}
    student_present_list = []
    student_absent_list = []
    date_list = []
    month_list = []

    for attendance in filtered_attendance:
        date_str = attendance.date.strftime('%Y-%m-%d')
        day = attendance.date.day
        month = attendance.date.strftime('%m')  # Lấy tháng dưới dạng số

        if date_str not in date_list:
            date_list.append(day)  # Lưu ngày
            month_list.append(month)  # Lưu tháng cho mỗi ngày

        present_count = AttendanceReport.objects.filter(attendance=attendance, status=True).count()
        absent_count = AttendanceReport.objects.filter(attendance=attendance, status=False).count()

        student_present_list.append(present_count)
        student_absent_list.append(absent_count)

    for course in courses:
        course_name = course.name  # Lấy tên khóa học
        # Khởi tạo
        if course_name not in course_student_counts:
            course_student_counts[course_name] = {}

        # Lấy danh sách lớp
        subjects_in_course = Subject.objects.filter(course=course)

        for subject in subjects_in_course:
            # Lấy số lượng học sinh trong lớp
            total_students_in_subject = Student.objects.filter(
                course=course, session=subject
            ).count()

            # Lưu số học sinh cho từng lớp
            course_student_counts[course_name][subject.name] = total_students_in_subject

    for subject in subjects:
        # Lấy danh sách lớp học
        subject_list.append(subject.name)

        # Lấy tất cả các bản ghi điểm danh trong lớp theo ngày đã chọn
        attendances = Attendance.objects.filter(
            subject=subject, date=selected_date  # Sử dụng ngày được chọn
        )

        # Số học sinh có mặt trong lớp vào ngày được chọn
        present_count = AttendanceReport.objects.filter(
            attendance__in=attendances, status=True  # Học sinh có mặt
        ).count()

        # Số học sinh vắng mặt trong lớp vào ngày được chọn
        absent_count = AttendanceReport.objects.filter(
            attendance__in=attendances, status=False  # Học sinh vắng mặt
        ).count()

        # Thêm số lượng học sinh đã điểm danh và vắng mặt vào danh sách
        student_present_today_list.append(present_count)
        student_absent_today_list.append(absent_count)

    # Lấy số lớp học có lịch trong ngày được chọn
    classes_with_schedule_today = (
        TeachingSchedule.objects.filter(schedule_date=selected_date)
        .values("class_name")
        .distinct()
        .count()
    )

    total_classes = subjects.count()

    context = {
        "page_title": "Bảng điều khiển ADMIN",
        "total_students": total_students,
        "total_staff": total_staff,
        "total_course": total_course,
        "subjects": subjects, #tên lớp
        "courses": courses,# tên trường
        "total_subject": total_subject,
        "subject_list": json.dumps(subject_list),
        "total_teaching": total_teaching,
        "student_absent_today_list": student_absent_today_list,  # Danh sách số học sinh vắng mặt theo lớp
        "student_present_today_list": student_present_today_list,  # Danh sách số học sinh có mặt theo lớp
        "total_classes": total_classes,
        "classes_with_schedule_today": classes_with_schedule_today,  # Số lớp có lịch học hôm nay
        "course_student_counts": json.dumps(course_student_counts),
        "student_present_list": json.dumps(student_present_list),
        "student_absent_list": json.dumps(student_absent_list),
        "date_list": json.dumps(date_list),
        "month_list": json.dumps(month_list), 
    }
    print(context)
    return render(request, "hod_template/home_content.html", context)


def filter_data_by_date(request):
    selected_date = request.GET.get("date", timezone.now().date())
    print(f"ngày được họn: {selected_date}")

    # Lọc dữ liệu học sinh điểm danh theo ngày đã chọn
    subjects = Subject.objects.all()

    student_present_today_list = []
    student_absent_today_list = []

    for subject in subjects:
        # Lấy tất cả các bản ghi điểm danh trong lớp theo ngày
        attendances = Attendance.objects.filter(subject=subject, date=selected_date)
        print(f"bản ghi điểm danh trong lớp theo ngày: {attendances}")

        # Số học sinh có mặt trong lớp vào ngày được chọn
        present_count = AttendanceReport.objects.filter(
            attendance__in=attendances, status=True  # Học sinh có mặt
        ).count()

        # Số học sinh vắng mặt trong lớp vào ngày được chọn
        absent_count = AttendanceReport.objects.filter(
            attendance__in=attendances, status=False  # Học sinh vắng mặt
        ).count()

        student_present_today_list.append(present_count)
        student_absent_today_list.append(absent_count)

    # Lấy số lớp học có lịch hôm nay
    classes_with_schedule_today = (
        TeachingSchedule.objects.filter(schedule_date=selected_date)
        .values("class_name")
        .distinct()
        .count()
    )

    total_classes = subjects.count()

    # Làm sạch dữ liệu trước khi trả về
    student_present_today_list = list(map(int, student_present_today_list))
    student_absent_today_list = list(map(int, student_absent_today_list))

    return JsonResponse(
        {
            "student_present_today_list": student_present_today_list,
            "student_absent_today_list": student_absent_today_list,
            "classes_with_schedule_today": classes_with_schedule_today,
            "total_classes": total_classes,
        }
    )


def add_staff(request):
    form = StaffForm(
        request.POST or None, request.FILES or None, user_type="staff"
    )  # Pass user_type to the form
    context = {"form": form, "page_title": "Thêm giáo viên", "user_type": "staff"}

    if request.method == "POST":
        if form.is_valid():
            first_name = form.cleaned_data.get("first_name")
            last_name = form.cleaned_data.get("last_name")
            address = form.cleaned_data.get("address")
            email = form.cleaned_data.get("email")
            gender = form.cleaned_data.get("gender")
            password = form.cleaned_data.get("password")

            passport = request.FILES.get("profile_pic")
            fs = FileSystemStorage()
            filename = fs.save(passport.name, passport)
            passport_url = fs.url(filename)
            try:
                user = CustomUser.objects.create_user(
                    email=email,
                    password=password,
                    user_type=2,
                    first_name=first_name,
                    last_name=last_name,
                    profile_pic=passport_url,
                )
                user.gender = gender
                user.address = address

                user.save()
                messages.success(request, "Đã thêm thành công")
                return redirect(reverse("manage_staff"))

            except Exception as e:
                messages.error(request, "Could Not Add " + str(e))
        else:
            messages.error(request, "Please fulfil all requirements")

    return render(request, "hod_template/add_staff_template.html", context)


def add_staff_csv(request):
    context = {"page_title": "Thêm nhiều giáo viên"}

    if request.method == "POST":
        excel_file = request.FILES.get("file")

        # Kiểm tra nếu không có file hoặc file trống
        if not excel_file:
            messages.error(request, "Vui lòng tải lên file Excel hợp lệ.")
            return redirect(reverse("add_staff"))

        # Kiểm tra định dạng file
        if not excel_file.name.endswith(".xlsx"):
            messages.error(
                request,
                "Định dạng file không hợp lệ. Vui lòng tải lên file Excel (.xlsx).",
            )
            return redirect(reverse("add_staff"))

        try:
            # Đọc file Excel
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active  # Lấy sheet đầu tiên

            # Kiểm tra số cột trong file Excel
            expected_columns = [
                "ho",
                "ten",
                "dia_chi",
                "email",
                "gioi_tinh",
                "mat_khau",
            ]
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if header != expected_columns:
                messages.error(
                    request,
                    "Định dạng file không đúng. Vui lòng đảm bảo file có các cột: ho, ten, dia_chi, email, gioi_tinh, mat_khau.",
                )
                return redirect(reverse("add_staff"))

            # Đường dẫn đến ảnh mặc định
            default_image_path = os.path.join(
                settings.MEDIA_ROOT, "image_default", "bdulogo.jpg"
            )
            media_image_path = os.path.join(settings.MEDIA_ROOT, "profile_pics")

            # Tạo thư mục chứa ảnh đại diện nếu chưa tồn tại
            if not os.path.exists(media_image_path):
                os.makedirs(media_image_path)

            # Copy ảnh mặc định ra ngoài thư mục media/profile_pics nếu chưa có
            default_profile_pic = os.path.join(media_image_path, "bdulogo.jpg")
            if not os.path.exists(default_profile_pic):
                try:
                    shutil.copy(default_image_path, default_profile_pic)
                except Exception as e:
                    messages.error(
                        request, f"Không thể sao chép ảnh đại diện mặc định: {str(e)}"
                    )
                    return redirect(reverse("add_staff"))

            # Lặp qua từng dòng trong Excel
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                ho = row[0]  # Họ
                ten = row[1]  # Tên
                dia_chi = row[2]  # Địa chỉ
                email = row[3]  # Email
                gioi_tinh = row[4]  # Giới tính
                mat_khau = str(row[5])  # Mật khẩu

                # Kiểm tra nếu dữ liệu dòng có vấn đề
                if not all([ho, ten, dia_chi, email, gioi_tinh, mat_khau]):
                    messages.error(
                        request,
                        f"Dữ liệu dòng {idx + 2} không hợp lệ. Vui lòng kiểm tra.",
                    )
                    continue

                # Kiểm tra định dạng email
                if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
                    messages.error(
                        request, f"Định dạng email không hợp lệ ở dòng {idx + 2}."
                    )
                    continue

                # Đặt ảnh đại diện mặc định
                anh_dai_dien_url = "/media/profile_pics/bdulogo.jpg"

                # Tạo tài khoản giáo viên
                try:
                    user = CustomUser.objects.create_user(
                        email=email,
                        password=mat_khau,
                        user_type=2,
                        first_name=ten,
                        last_name=ho,
                        profile_pic=anh_dai_dien_url,
                    )
                    user.gender = gioi_tinh
                    user.address = dia_chi
                    user.save()

                except Exception as e:
                    messages.error(
                        request, f"Không thể thêm giáo viên {email}: {str(e)}"
                    )
                    continue

            messages.success(request, "Đã thêm danh sách giáo viên thành công.")
            return redirect(reverse("manage_staff"))

        except Exception as e:
            messages.error(request, f"Đã xảy ra lỗi khi xử lý file Excel: {str(e)}")
            return redirect(reverse("add_staff"))

    return render(request, "hod_template/add_staff_template.html", context)


def add_student(request):
    student_form = StudentForm(
        request.POST or None, request.FILES or None, user_type="student"
    )
    context = {
        "form": student_form,
        "page_title": "Thêm học sinh",
        "user_type": "student",
    }
    if request.method == "POST":
        if student_form.is_valid():
            student_id = student_form.cleaned_data.get("student_id")

            # Kiểm tra xem student_id có phải là số hay không
            # if not student_id.isdigit():
            #     messages.error(request, "Mã học sinh phải là số")
            #     return render(
            #         request, "hod_template/add_student_template.html", context
            #     )

            # Các dữ liệu khác sau khi kiểm tra student_id
            first_name = student_form.cleaned_data.get("first_name")
            last_name = student_form.cleaned_data.get("last_name")
            address = student_form.cleaned_data.get("address")
            gender = student_form.cleaned_data.get("gender")
            password = student_form.cleaned_data.get("password")
            course = student_form.cleaned_data.get("course")
            session = student_form.cleaned_data.get("session")

            try:
                user = CustomUser.objects.create_user(
                    student_id=student_id,
                    password=password,
                    user_type=3,
                    first_name=first_name,
                    last_name=last_name,
                )

                user.gender = gender
                user.address = address
                user.student.session = session
                user.student.course = course
                user.save()

                # Tạo ảnh QR
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=10,
                    border=4,
                )
                qr.add_data(student_id)
                qr.make(fit=True)

                img = qr.make_image(fill="black", back_color="white")
                buffer = BytesIO()
                img.save(buffer, format="PNG")
                qr_image_name = f"{student_id}_qr.png"

                qr_dir = os.path.join(settings.MEDIA_ROOT, "QR")
                if not os.path.exists(qr_dir):
                    os.makedirs(qr_dir)

                qr_path = os.path.join(qr_dir, qr_image_name)
                with open(qr_path, "wb") as f:
                    f.write(buffer.getvalue())

                user.qr_code = os.path.join("QR", qr_image_name)
                user.save()
                messages.success(request, "Đã thêm thành công")
                return redirect(reverse("manage_student"))
            except Exception as e:
                messages.error(request, "Could Not Add: " + str(e))
        else:
            messages.error(request, "Could Not Add: ")
    return render(request, "hod_template/add_student_template.html", context)


def add_course(request):
    form = CourseForm(request.POST or None)
    context = {"form": form, "page_title": "Thêm tên trường"}
    if request.method == "POST":
        if form.is_valid():
            name = form.cleaned_data.get("name")
            try:
                course = Course()
                course.name = name
                course.save()
                messages.success(request, "Đã thêm thành công")
                return redirect(reverse("manage_course"))
            except:
                messages.error(request, "Could Not Add")
        else:
            messages.error(request, "Could Not Add")
    return render(request, "hod_template/add_course_template.html", context)


def add_subject(request):
    form = SubjectForm(request.POST or None)
    context = {"form": form, "page_title": "Thêm lớp"}
    if request.method == "POST":
        if form.is_valid():
            name = form.cleaned_data.get("name")
            course = form.cleaned_data.get("course")
            staff = form.cleaned_data.get("staff")
            try:
                subject = Subject()
                subject.name = name
                # subject.staff = staff
                subject.course = course
                subject.save()
                messages.success(request, "Đã thêm thành công")
                return redirect(reverse("manage_subject"))

            except Exception as e:
                messages.error(request, "Could Not Add " + str(e))
        else:
            messages.error(request, "Fill Form Properly")

    return render(request, "hod_template/add_subject_template.html", context)


def manage_staff(request):
    allStaff = CustomUser.objects.filter(user_type=2)
    paginator = Paginator(allStaff, 10)  # Mỗi trang có 10 lịch dạy
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {"allStaff": allStaff,"page_obj": page_obj,  "page_title": "Quản lý giáo viên"}
    return render(request, "hod_template/manage_staff.html", context)


from django.db.models import Prefetch


def manage_student(request):
    # Retrieve all students and prefetch related TeachingSchedule to get class name (subject)
    students = Student.objects.select_related("admin", "course").prefetch_related(
        Prefetch(
            "course__teachingschedule_set",
            queryset=TeachingSchedule.objects.all(),
            to_attr="teaching_schedules",
        )
    )

    context = {"students": students, "page_title": "Quản lý học sinh"}
    return render(request, "hod_template/manage_student.html", context)


def manage_course(request):
    courses = Course.objects.all()
    context = {"courses": courses, "page_title": "Quản lý trường "}
    return render(request, "hod_template/manage_course.html", context)


def manage_subject(request):
    subjects = Subject.objects.all()
    context = {"subjects": subjects, "page_title": "Quản lý lớp"}
    return render(request, "hod_template/manage_subject.html", context)


def edit_staff(request, staff_id):
    # Lấy thông tin staff dựa trên staff_id
    staff = get_object_or_404(Staff, id=staff_id)

    # Khởi tạo form với instance của Staff
    form = StaffForm(
        request.POST or None,
        request.FILES or None,
        instance=staff.admin,
        user_type="staff",
    )

    context = {
        "form": form,
        "staff_id": staff_id,
        "page_title": "Chỉnh sửa thông tin giáo viên",
        "user_type": "staff",
    }

    if request.method == "POST" and form.is_valid():
        try:
            user = staff.admin
            # Cập nhật thông tin user từ form
            user.first_name = form.cleaned_data.get("first_name")
            user.last_name = form.cleaned_data.get("last_name")
            user.gender = form.cleaned_data.get("gender")
            user.address = form.cleaned_data.get("address")

            # Cập nhật email nếu là Staff
            if user.user_type == "2":  # Staff
                user.email = form.cleaned_data.get("email")
                user.student_id = None
            else:
                user.email = None

            # Nếu có mật khẩu mới
            password = form.cleaned_data.get("password")
            if password:
                user.set_password(password)

            # Nếu có hình ảnh mới
            profile_pic = request.FILES.get("profile_pic")
            if profile_pic:
                fs = FileSystemStorage()
                filename = fs.save(profile_pic.name, profile_pic)
                user.profile_pic = fs.url(filename)

            # Lưu thông tin user và staff
            user.save()
            staff.save()

            messages.success(request, "Cập nhật thành công")
            # return redirect(reverse("edit_staff", args=[staff_id]))
            return redirect(reverse("manage_staff"))
        except Exception as e:
            messages.error(request, "Không thể cập nhật: " + str(e))
    else:
        messages.error(request, "Vui lòng điền đầy đủ thông tin vào biểu mẫu")

    return render(request, "hod_template/edit_staff_template.html", context)

def edit_student(request, student_id):
    student = get_object_or_404(Student, admin_id=student_id)
    user = student.admin  # Truy cập đối tượng CustomUser thông qua Student

    # Tạo form với instance là user
    form = StudentForm(
        request.POST or None,
        request.FILES or None,
        instance=user,
        user_type="student",
    )

    # Điền thông tin vào context
    context = {
        "form": form,
        "student_id": student_id,
        "page_title": "Chỉnh sửa thông tin học sinh",
        "user_type": "student",
    }

    if request.method == "POST" and form.is_valid():
        try:
            # Cập nhật thông tin người dùng
            user.student_id = form.cleaned_data.get("student_id")
            user.first_name = form.cleaned_data.get("first_name")
            user.last_name = form.cleaned_data.get("last_name")
            user.gender = form.cleaned_data.get("gender")
            user.address = form.cleaned_data.get("address")

            # Cập nhật student_id nếu là Học sinh
            if user.user_type == "3":  # Học sinh
                user.student_id = form.cleaned_data.get("student_id")
                user.email = None

            # Cập nhật mật khẩu nếu có mật khẩu mới
            password = form.cleaned_data.get("password")
            if password:
                user.set_password(password)

            # Lưu ảnh đại diện nếu có ảnh mới được tải lên
            profile_pic = request.FILES.get("profile_pic")
            if profile_pic:
                fs = FileSystemStorage()
                filename = fs.save(profile_pic.name, profile_pic)
                user.profile_pic = fs.url(filename)

            # Cập nhật thông tin khóa học và lớp học
            student.course = form.cleaned_data.get("course")
            student.session = form.cleaned_data.get("session")

            # Tạo và lưu QR code
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(user.student_id)  # Sử dụng student_id từ user
            qr.make(fit=True)

            img = qr.make_image(fill="black", back_color="white")
            buffer = BytesIO()
            img.save(buffer, format="PNG")
            qr_image_name = f"{user.student_id}_qr.png"

            # Đảm bảo thư mục QR code tồn tại
            qr_dir = os.path.join(settings.MEDIA_ROOT, "QR")
            if not os.path.exists(qr_dir):
                os.makedirs(qr_dir)

            qr_path = os.path.join(qr_dir, qr_image_name)

            # Lưu QR code vào thư mục QR
            with open(qr_path, "wb") as f:
                f.write(buffer.getvalue())

            # Cập nhật đường dẫn QR code
            user.qr_code = os.path.relpath(qr_path, settings.MEDIA_ROOT)

            # Lưu thông tin người dùng và học sinh
            user.save()
            student.save()

            messages.success(request, "Cập nhật thành công")
            # return redirect(reverse("edit_student", args=[student_id]))
            return redirect(reverse("manage_student"))
        except Exception as e:
            messages.error(request, "Không thể cập nhật: " + str(e))
    else:
        messages.error(request, "Vui lòng điền đầy đủ thông tin vào biểu mẫu!")

    return render(request, "hod_template/edit_student_template.html", context)


def edit_course(request, course_id):
    instance = get_object_or_404(Course, id=course_id)
    form = CourseForm(request.POST or None, instance=instance)
    context = {
        "form": form,
        "course_id": course_id,
        "page_title": "Chỉnh sửa thông tin khóa học",
    }
    if request.method == "POST":
        if form.is_valid():
            name = form.cleaned_data.get("name")
            try:
                course = Course.objects.get(id=course_id)
                course.name = name
                course.save()
                messages.success(request, "Cập nhật thành công")
                # return redirect(reverse("edit_course", args=[course_id]))
                return redirect(reverse("manage_course"))
            except:
                messages.error(request, "Không thể cập nhật")
        else:
            messages.error(request, "Không thể cập nhật")

    return render(request, "hod_template/edit_course_template.html", context)


def edit_subject(request, subject_id):
    instance = get_object_or_404(Subject, id=subject_id)
    form = SubjectForm(request.POST or None, instance=instance)
    context = {
        "form": form,
        "subject_id": subject_id,
        "page_title": "Chỉnh sửa thông tin lớp",
    }
    if request.method == "POST":
        if form.is_valid():
            name = form.cleaned_data.get("name")
            course = form.cleaned_data.get("course")
            staff = form.cleaned_data.get("staff")
            try:
                subject = Subject.objects.get(id=subject_id)
                subject.name = name
                subject.staff = staff
                subject.course = course
                subject.save()
                messages.success(request, "Cập nhật thành công")
                # return redirect(reverse("edit_subject", args=[subject_id]))
                return redirect(reverse("manage_subject"))
            except Exception as e:
                messages.error(request, "Không thể thêm" + str(e))
        else:
            messages.error(request, "Điền biểu mấu đúng cách")
    return render(request, "hod_template/edit_subject_template.html", context)


def add_session(request):
    if request.method == "POST":
        form = TeachingScheduleForm(request.POST)
        if form.is_valid():
            try:
                teaching_schedule = form.save()
                # Lấy thông tin email một lần, tránh gọi nhiều lần
                
                schedules = TeachingSchedule.objects.filter(staff=teaching_schedule.staff)
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Lịch dạy"
                worksheet.append(
                    ["Tên giáo viên", "Trường", "Lớp", "Ngày dạy", "Thời gian"]
                ) 
                for schedule in schedules:
                    # add to xlsx file 
                    if schedule.schedule_date < timezone.now().date():
                        continue
                    worksheet.append(
                        [
                            schedule.staff.admin.last_name + " " + schedule.staff.admin.first_name,
                            schedule.school_name.name,
                            schedule.class_name.name,
                            schedule.schedule_date.strftime("%d/%m/%Y"),
                            schedule.start_time.strftime("%H:%M"),
                        ]
                    )
                excel_file = BytesIO()
                workbook.save(excel_file)
                excel_file.seek(0)
                admin = teaching_schedule.staff.admin
                teacher_email = admin.email
                subject = "Thông báo lịch dạy mới"
                message = (
                    f"Xin chào Thầy/Cô {admin.last_name} {admin.first_name},\n\n"
                    f"Quý Thầy/Cô vừa được phân công lịch giảng dạy mới với thông tin chi tiết như sau:\n"
                    f"- Ngày giảng dạy: {teaching_schedule.schedule_date.strftime('%d/%m/%Y')}\n"
                    f"- Thời gian: {teaching_schedule.start_time.strftime('%H:%M')}\n"
                    f"- Lớp học: {teaching_schedule.class_name}\n"
                    f"- Trường: {teaching_schedule.school_name}\n\n"
                    "Kính mong Quý Thầy/Cô kiểm tra lại toàn bộ lịch dạy trong file đính kèm và có những chuẩn bị cần thiết để buổi giảng dạy diễn ra thành công./.\n\n"
                    "Trân trọng,\n \n"
                    "HỆ THỐNG QUẢN LÝ ĐIỂM DANH BDU CÀ MAU"

                )
                # Gửi email
                email = EmailMessage(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [teacher_email],
                )

                # Đính kèm file Excel
                email.attach(
                    f"Lịch dạy giảng viên {teaching_schedule.staff.admin.last_name} {teaching_schedule.staff.admin.first_name}".upper() + ".xlsx",
                    excel_file.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                email.send()    
            #     send_mail(
            #         subject,
            #         message,
            #         settings.DEFAULT_FROM_EMAIL,
            #         [teacher_email],
            #         fail_silently=False,
            #     )
                # Hiển thị thông báo thành công
                messages.success(
                    request,
                    f"Tạo lịch dạy thành công và thông báo đã được gửi gmail của giáo viên {admin.last_name} {admin.first_name}",
                )
                return redirect(reverse("manage_session"))

            except Exception as e:
                messages.error(request, f"Không thể thêm lịch dạy: {e}")
        else:
            messages.error(request, "Điền biểu mẫu đúng cách")
    else:
        form = TeachingScheduleForm()

    context = {"form": form, "page_title": "Thêm lịch dạy"}
    return render(request, "hod_template/add_session_template.html", context)


def load_classes(request):
    school_id = request.GET.get("school_id")
    classes = Subject.objects.filter(course_id=school_id).order_by("name")
    return JsonResponse(list(classes.values("id", "name")), safe=False)


def load_classes_student(request):
    course_id = request.GET.get("course_id")
    subjects = Subject.objects.filter(course_id=course_id).order_by("name")
    return JsonResponse(list(subjects.values("id", "name")), safe=False)


def manage_session(request):
    sessions = TeachingSchedule.objects.all()
    paginator = Paginator(sessions, 10)  # Mỗi trang có 10 lịch dạy
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {"sessions": sessions,'page_obj': page_obj, "page_title": "Quản lý lịch dạy"}
    return render(request, "hod_template/manage_session.html", context)

def delete_schedules(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)  # Lấy danh sách IDs từ yêu cầu JSON
            ids = data.get('ids', [])

            if ids and all(ids):  # Kiểm tra xem danh sách IDs có giá trị
                TeachingSchedule.objects.filter(id__in=ids).delete()  # Xóa các lịch dạy tương ứng
                messages.success(request, "Xóa thành công lịch dạy đã chọn.")
                return JsonResponse({'status': 'success'}, status=200)
            else:
                return JsonResponse({'status': 'error', 'message': 'No valid IDs provided'}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON format'}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


def edit_session(request, session_id):
    # Fetch the TeachingSchedule instance based on session_id
    instance = get_object_or_404(TeachingSchedule, id=session_id)

    if request.method == "POST":
        form = TeachingScheduleForm(
            request.POST, instance=instance
        )  # Initialize form with POST data and instance
        if form.is_valid():
            try:
                updated_teaching_schedule = form.save()
                schedules = TeachingSchedule.objects.filter(staff=updated_teaching_schedule.staff)
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Lịch dạy"
                worksheet.append(
                    ["Tên giáo viên", "Trường", "Lớp", "Ngày dạy", "Thời gian"]
                ) 
                for schedule in schedules:
                    # add to xlsx file 
                    if schedule.schedule_date < timezone.now().date():
                        continue
                    worksheet.append(
                        [
                            schedule.staff.admin.last_name + " " + schedule.staff.admin.first_name,
                            schedule.school_name.name,
                            schedule.class_name.name,
                            schedule.schedule_date.strftime("%d/%m/%Y"),
                            schedule.start_time.strftime("%H:%M"),
                        ]
                    )
                excel_file = BytesIO()
                workbook.save(excel_file)
                excel_file.seek(0)
                # Send email notification to the teacher
                admin = updated_teaching_schedule.staff.admin
                teacher_email = admin.email
                subject = "Thông báo cập nhật lịch dạy"
                message = (
                    f"Xin chào Thầy/Cô {admin.last_name} {admin.first_name},\n\n"
                    f"Quý Thầy/Cô vừa được cập nhật lịch giảng dạy với thông tin chi tiết như sau:\n"
                    f"- Ngày giảng dạy: {updated_teaching_schedule.schedule_date.strftime('%d/%m/%Y')}\n"
                    f"- Thời gian: {updated_teaching_schedule.start_time.strftime('%H:%M')}\n"
                    f"- Lớp học: {updated_teaching_schedule.class_name}\n"
                    f"- Trường: {updated_teaching_schedule.school_name}\n\n"
                    "Kính mong Quý Thầy/Cô kiểm tra lại toàn bộ lịch dạy trong file đính kèm và có những chuẩn bị cần thiết để buổi giảng dạy diễn ra thành công./.\n\n"
                    "Trân trọng,\n \n"
                    "HỆ THỐNG QUẢN LÝ ĐIỂM DANH BDU CÀ MAU"
                )

                # Gửi email
                email = EmailMessage(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [teacher_email],
                )

                # Đính kèm file Excel
                email.attach(
                    f"Lịch dạy giảng viên {updated_teaching_schedule.staff.admin.last_name} {updated_teaching_schedule.staff.admin.first_name}".upper() + ".xlsx",
                    excel_file.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                email.send() 

                messages.success(
                    request,
                    f"Cập nhật thành công lịch dạy và thông báo đã được gửi đến gamil của giáo viên {admin.last_name} {admin.first_name}.",
                )
                # return redirect(reverse("edit_session", args=[session_id]))
                return redirect(reverse("manage_session"))
            except Exception as e:
                messages.error(
                    request, "Hệ thống không thể cập nhật thông tin: " + str(e)
                )
        else:
            messages.error(request, "Đã gửi biểu mẫu không hợp lệ")
    else:
        form = TeachingScheduleForm(instance=instance)  # Initialize form with instance

    context = {
        "form": form,
        "session_id": session_id,
        "page_title": "Chỉnh sửa lịch dạy",
    }
    return render(request, "hod_template/edit_session_template.html", context)


@csrf_exempt
def check_email_availability(request):
    email = request.POST.get("email")
    try:
        user = CustomUser.objects.filter(email=email).exists()
        if user:
            return HttpResponse(True)
        return HttpResponse(False)
    except Exception as e:
        return HttpResponse(False)


@csrf_exempt
def student_feedback_message(request):
    if request.method != "POST":
        feedbacks = FeedbackStudent.objects.all()
        context = {
            "feedbacks": feedbacks,
            "page_title": "Tin nhắn phản hồi của học sinh",
        }
        return render(request, "hod_template/student_feedback_template.html", context)
    else:
        feedback_id = request.POST.get("id")
        try:
            feedback = get_object_or_404(FeedbackStudent, id=feedback_id)
            reply = request.POST.get("reply")
            feedback.reply = reply
            feedback.save()
            return HttpResponse(True)
        except Exception as e:
            return HttpResponse(False)


@csrf_exempt
def staff_feedback_message(request):
    if request.method != "POST":
        feedbacks = FeedbackStaff.objects.all()
        context = {
            "feedbacks": feedbacks,
            "page_title": "Tin nhắn phản hồi của giáo viên",
        }
        return render(request, "hod_template/staff_feedback_template.html", context)
    else:
        feedback_id = request.POST.get("id")
        try:
            feedback = get_object_or_404(FeedbackStaff, id=feedback_id)
            reply = request.POST.get("reply")
            feedback.reply = reply
            feedback.save()
            return HttpResponse(True)
        except Exception as e:
            return HttpResponse(False)


@csrf_exempt
def view_staff_leave(request):
    if request.method != "POST":
        allLeave = LeaveReportStaff.objects.all()
        context = {"allLeave": allLeave, "page_title": "Đơn xin nghỉ phép từ giáo viên"}
        return render(request, "hod_template/staff_leave_view.html", context)
    else:
        id = request.POST.get("id")
        status = request.POST.get("status")
        if status == "1":
            status = 1
        else:
            status = -1
        try:
            leave = get_object_or_404(LeaveReportStaff, id=id)
            leave.status = status
            leave.save()
            return HttpResponse(True)
        except Exception as e:
            return False


@csrf_exempt
def view_student_leave(request):
    if request.method != "POST":
        allLeave = LeaveReportStudent.objects.all()
        context = {"allLeave": allLeave, "page_title": "Đơn xin nghỉ phép từ học sinh"}
        return render(request, "hod_template/student_leave_view.html", context)
    else:
        id = request.POST.get("id")
        status = request.POST.get("status")
        if status == "1":
            status = 1
        else:
            status = -1
        try:
            leave = get_object_or_404(LeaveReportStudent, id=id)
            leave.status = status
            leave.save()
            return HttpResponse(True)
        except Exception as e:
            return False


def admin_view_attendance(request):
    subjects = Subject.objects.all()
    sessions = Course.objects.all()
    context = {
        "subjects": subjects,
        "sessions": sessions,
        "page_title": "Xem điểm danh",
    }

    return render(request, "hod_template/admin_view_attendance.html", context)


def get_subjects_by_session(request):
    session_id = request.GET.get("session_id")
    subjects = Subject.objects.filter(course_id=session_id)  # Lọc lớp theo trường
    subject_list = [{"id": subject.id, "name": subject.name} for subject in subjects]

    return JsonResponse({"subjects": subject_list})


@csrf_exempt
def get_admin_attendance(request):
    subject_id = request.POST.get("subject")
    session_id = request.POST.get("session")
    attendance_date_id = request.POST.get("attendance_date_id")
    start_date = request.POST.get("start_date")
    end_date = request.POST.get("end_date")

    if start_date:
        try:
            if "/" in start_date:
                start_date = datetime.strptime(start_date, "%d/%m/%Y").strftime(
                    "%Y-%m-%d"
                )
        except ValueError as e:
            return JsonResponse(
                {"error": "Định dạng ngày bắt đầu không hợp lệ"}, status=400
            )

    if end_date:
        try:
            if "/" in end_date:
                end_date = datetime.strptime(end_date, "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError as e:
            return JsonResponse(
                {"error": "Định dạng ngày kết thúc không hợp lệ"}, status=400
            )

    try:
        subject = get_object_or_404(Subject, id=subject_id)
        session = get_object_or_404(Course, id=session_id)

        if start_date and end_date:
            attendance_reports = AttendanceReport.objects.filter(
                attendance__subject=subject,
                attendance__session=session,
                attendance__date__range=(start_date, end_date),
            )
        elif attendance_date_id:
            attendance = get_object_or_404(
                Attendance, id=attendance_date_id, session=session
            )
            attendance_reports = AttendanceReport.objects.filter(attendance=attendance)
        else:
            return JsonResponse({"error": "Thiếu thông tin cần thiết"}, status=400)

        # Chuẩn bị dữ liệu JSON với thông tin thêm về lớp và trường
        json_data = [
            {
                "student_id": report.student.id,
                "student_name": str(report.student),
                "status": report.status,
                "attendance_date": report.attendance.date.strftime("%d/%m/%Y"),
                "school_name": report.attendance.session.name,  # Tên trường
                "class_name": report.attendance.subject.name,  # Tên lớp
            }
            for report in attendance_reports
        ]
        print(json_data)
        return JsonResponse(json_data, safe=False)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@csrf_exempt
def delete_attendance(request):
    if request.method == "POST":
        subject_id = request.POST.get("subject_id")
        session_id = request.POST.get("session_id")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        attendance_date = request.POST.get("attendance_date")

        print("subject_id: ", subject_id)
        print("session_id: ", session_id)
        print("start_date: ", start_date)
        print("end_date: ", end_date)
        print("attendance_date: ", attendance_date)

        try:
            # Chuyển đổi định dạng ngày từ DD/MM/YYYY sang YYYY-MM-DD
            def convert_date(date_str, format_str="%d/%m/%Y"):
                try:
                    return datetime.strptime(date_str, format_str).strftime("%Y-%m-%d")
                except ValueError:
                    return None

            def convert_date_range(date_str, format_str="%Y-%m-%d"):
                try:
                    return datetime.strptime(date_str, format_str).strftime("%Y-%m-%d")
                except ValueError:
                    return None

            formatted_attendance_date = convert_date(attendance_date)
            formatted_start_date = (
                convert_date_range(start_date) if start_date else None
            )
            formatted_end_date = convert_date_range(end_date) if end_date else None

            print(f"Formatted Attendance Date: {formatted_attendance_date}")
            print(f"Formatted Start Date: {formatted_start_date}")
            print(f"Formatted End Date: {formatted_end_date}")

            if formatted_start_date and formatted_end_date:
                # Xóa dữ liệu theo khoảng thời gian
                attendance_records = Attendance.objects.filter(
                    subject_id=subject_id,
                    session_id=session_id,
                    date__range=[formatted_start_date, formatted_end_date],
                )
                print(
                    "Khoảng thời gian:", formatted_start_date, "đến", formatted_end_date
                )
                print("Dữ liệu tìm thấy:", attendance_records)

                # Xóa các bản ghi liên quan trong AttendanceReport
                AttendanceReport.objects.filter(
                    attendance_id__in=attendance_records.values_list("id", flat=True)
                ).delete()
                # Xóa các bản ghi trong Attendance
                attendance_records.delete()

            elif formatted_attendance_date:
                # Xóa dữ liệu theo ngày điểm danh
                attendance_records = Attendance.objects.filter(
                    subject_id=subject_id,
                    session_id=session_id,
                    date=formatted_attendance_date,
                )
                print("Dữ liệu theo ngày điểm danh:", attendance_records)
                # Xóa các bản ghi liên quan trong AttendanceReport
                AttendanceReport.objects.filter(
                    attendance_id__in=attendance_records.values_list("id", flat=True)
                ).delete()
                # Xóa các bản ghi trong Attendance
                attendance_records.delete()

            else:
                return JsonResponse(
                    {"status": "error", "message": "Không có điều kiện xóa dữ liệu!"}
                )

            return JsonResponse(
                {"status": "success", "message": "Dữ liệu đã được xóa thành công!"}
            )
        except Exception as e:
            return JsonResponse(
                {"status": "error", "message": f"Có lỗi xảy ra: {str(e)}"}
            )


def admin_view_profile(request):
    admin = get_object_or_404(Admin, admin=request.user)
    form = AdminForm(
        request.POST or None, request.FILES or None, instance=admin, user_type="admin"
    )
    context = {"form": form, "page_title": "Chỉnh sửa thông tin", "user_type": "admin"}
    if request.method == "POST":
        try:
            if form.is_valid():
                first_name = form.cleaned_data.get("first_name")
                last_name = form.cleaned_data.get("last_name")
                password = form.cleaned_data.get("password") or None
                passport = request.FILES.get("profile_pic") or None
                custom_user = admin.admin
                if password != None:
                    custom_user.set_password(password)
                if passport != None:
                    fs = FileSystemStorage()
                    filename = fs.save(passport.name, passport)
                    passport_url = fs.url(filename)
                    custom_user.profile_pic = passport_url
                custom_user.first_name = first_name
                custom_user.last_name = last_name
                custom_user.save()
                messages.success(request, "Đã cập nhật hồ sơ!")
                return redirect(reverse("admin_view_profile"))
            else:
                messages.error(request, "Dữ liệu được cung cấp không hợp lệ")
        except Exception as e:
            messages.error(request, "Đã xảy ra lỗi khi cập nhật hồ sơ " + str(e))
    return render(request, "hod_template/admin_view_profile.html", context)


def admin_notify_staff(request):
    staff = CustomUser.objects.filter(user_type=2)
    context = {"page_title": "Gửi thông báo đến giáo viên.", "allStaff": staff}
    return render(request, "hod_template/staff_notification.html", context)


def admin_notify_student(request):
    student = CustomUser.objects.filter(user_type=3)
    context = {"page_title": "Gửi thông báo đến học sinh", "students": student}
    return render(request, "hod_template/student_notification.html", context)


@csrf_exempt
def send_student_notification(request):
    id = request.POST.get("id")
    message = request.POST.get("message")
    student = get_object_or_404(Student, admin_id=id)
    try:
        url = "https://fcm.googleapis.com/fcm/send"
        body = {
            "notification": {
                "title": "Hệ thống quản lý học sinh",
                "body": message,
                "click_action": reverse("student_view_notification"),
                "icon": static("dist/img/AdminLTELogo.png"),
            },
            "to": student.admin.fcm_token,
        }
        headers = {
            "Authorization": "key=AAAA3Bm8j_M:APA91bElZlOLetwV696SoEtgzpJr2qbxBfxVBfDWFiopBWzfCfzQp2nRyC7_A2mlukZEHV4g1AmyC6P_HonvSkY2YyliKt5tT3fe_1lrKod2Daigzhb2xnYQMxUWjCAIQcUexAMPZePB",
            "Content-Type": "application/json",
        }
        data = requests.post(url, data=json.dumps(body), headers=headers)
        notification = NotificationStudent(student=student, message=message)
        notification.save()
        return HttpResponse("True")
    except Exception as e:
        return HttpResponse("False")


@csrf_exempt
def send_staff_notification(request):
    id = request.POST.get("id")
    message = request.POST.get("message")
    staff = get_object_or_404(Staff, admin_id=id)
    try:
        url = "https://fcm.googleapis.com/fcm/send"
        body = {
            "notification": {
                "title": "Hệ thống quản lý học sinh",
                "body": message,
                "click_action": reverse("staff_view_notification"),
                "icon": static("dist/img/AdminLTELogo.png"),
            },
            "to": staff.admin.fcm_token,
        }
        headers = {
            "Authorization": "key=AAAA3Bm8j_M:APA91bElZlOLetwV696SoEtgzpJr2qbxBfxVBfDWFiopBWzfCfzQp2nRyC7_A2mlukZEHV4g1AmyC6P_HonvSkY2YyliKt5tT3fe_1lrKod2Daigzhb2xnYQMxUWjCAIQcUexAMPZePB",
            "Content-Type": "application/json",
        }
        data = requests.post(url, data=json.dumps(body), headers=headers)
        notification = NotificationStaff(staff=staff, message=message)
        notification.save()
        return HttpResponse("True")
    except Exception as e:
        return HttpResponse("False")


def delete_staff(request, staff_id):
    try:
        staff = get_object_or_404(CustomUser, staff__id=staff_id)

        print("ma nhan vien: ", staff)
        # Xóa ảnh đại diện nếu có
        if staff.profile_pic:
            profile_pic_name = staff.profile_pic.name
            print("Đường dẫn ảnh đại diện:", profile_pic_name)

            # Loại bỏ '/media/' khỏi đường dẫn trước khi kiểm tra và xóa
            if profile_pic_name.startswith("/media/"):
                profile_pic_name = profile_pic_name[7:]  # Bỏ cả dấu '/'

            full_profile_pic_path = os.path.join(settings.MEDIA_ROOT, profile_pic_name)
            print("Đường dẫn đầy đủ ảnh đại diện:", full_profile_pic_path)

            # Xóa ảnh đại diện nếu tệp tồn tại
            if os.path.exists(full_profile_pic_path):
                default_storage.delete(full_profile_pic_path)
        # Xóa giáo viên
        staff.delete()
        messages.success(request, "Xóa giáo viên thành công!")
    except Exception as e:
        messages.error(request, f"Đã xảy ra lỗi khi xóa giáo viên: {str(e)}")

    return redirect(reverse("manage_staff"))


from django.core.files.storage import default_storage


@csrf_exempt
def delete_students(request):
    if request.method == "POST":
        data = json.loads(request.body)
        ids = data.get("ids", [])

        # Dùng transaction.atomic để đảm bảo tính toàn vẹn dữ liệu
        with transaction.atomic():
            try:

                success_messages = []
                error_messages = []
                # Lấy danh sách học sinh cần xóa
                students = CustomUser.objects.filter(id__in=ids)

                # Lưu danh sách các file cần xóa
                profile_pics_to_delete = []
                qr_codes_to_delete = []

                for student in students:
                    if student.profile_pic and default_storage.exists(
                        student.profile_pic.path
                    ):
                        profile_pics_to_delete.append(student.profile_pic.path)

                    if student.qr_code and default_storage.exists(student.qr_code.path):
                        qr_codes_to_delete.append(student.qr_code.path)

                # Xóa các học sinh trong database cùng lúc
                deleted_count, _ = students.delete()

                # Xóa các file profile_pic và qr_code nếu tồn tại
                for pic in profile_pics_to_delete:
                    default_storage.delete(pic)

                for qr in qr_codes_to_delete:
                    default_storage.delete(qr)

                success_messages.append(f"Xóa học sinh thành công!")
            except Exception as e:
                error_messages.append(f"Đã xảy ra lỗi khi xóa học sinh.")

        # Hiển thị thông báo sau khi hoàn tất tất cả các thao tác xóa
        if success_messages:
            messages.success(request, " | ".join(success_messages))
        if error_messages:
            messages.error(request, " | ".join(error_messages))

        return JsonResponse({"success": True})

    return JsonResponse({"success": False}, status=400)


def delete_student(request, student_id):
    try:
        student = get_object_or_404(CustomUser, id=student_id)

        print("xoa học sinh:", student)
        # Xóa ảnh đại diện nếu có
        if student.profile_pic:
            profile_pic_name = student.profile_pic.name
            print("Đường dẫn ảnh đại diện:", profile_pic_name)

            # Loại bỏ '/media/' khỏi đường dẫn trước khi kiểm tra và xóa
            if profile_pic_name.startswith("/media/"):
                profile_pic_name = profile_pic_name[7:]  # Bỏ cả dấu '/'

            full_profile_pic_path = os.path.join(settings.MEDIA_ROOT, profile_pic_name)
            print("Đường dẫn đầy đủ ảnh đại diện:", full_profile_pic_path)

            # Xóa ảnh đại diện nếu tệp tồn tại
            if os.path.exists(full_profile_pic_path):
                default_storage.delete(full_profile_pic_path)

        # Xóa mã QR nếu có
        if student.qr_code:
            qr_code_path = os.path.join(settings.MEDIA_ROOT, student.qr_code.name)
            if os.path.exists(qr_code_path):
                default_storage.delete(qr_code_path)

        # Xóa học sinh
        student.delete()
        messages.success(request, "Xóa học sinh thành công!")
    except Exception as e:
        messages.error(
            request, f"Đã xảy ra lỗi khi xóa học sinh ID {student_id}: {str(e)}"
        )

    return redirect(reverse("manage_student"))


# xóa khóa học
def delete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    try:
        course.delete()
        messages.success(request, "Xóa khóa học thành công!")
    except Exception as e:
        messages.error(request, f"Xin lỗi, không thể xóa khóa học. Lỗi: {str(e)}")
    return redirect(reverse("manage_course"))


# xóa môn học
def delete_subject(request, subject_id):
    try:
        subject = get_object_or_404(Subject, id=subject_id)
        subject.delete()
        messages.success(request, "Xóa môn học thành công!")
    except Exception as e:
        messages.error(request, f"Đã xảy ra lỗi khi xóa môn học: {str(e)}")
    return redirect(reverse("manage_subject"))


# xóa lịch học


def delete_session(request, session_id):
    session = get_object_or_404(TeachingSchedule, id=session_id)
    try:
        session.delete()
        messages.success(request, "Xóa lịch dạy thành công!")
    except Exception as e:
        messages.error(request, f"Xin lỗi, không thể xóa lịch dạy này. Lỗi: {str(e)}")
    return redirect(reverse("manage_session"))


from django.db import transaction


# đăng ký tài khoản tự động
def import_csv(request):
    if request.method == "POST" and "csv_file" in request.FILES:
        csv_file = request.FILES["csv_file"]
        success_message_shown = False
        try:
            import_csv_data(csv_file, request)
            if not success_message_shown:
                messages.success(
                    request, "Import CSV thành công, tất cả tài khoản đã được đăng ký!"
                )
                success_message_shown = True
        except Exception as e:
            messages.error(request, f"Đã có lỗi xảy ra: {str(e)}")
        return redirect("manage_student")
    else:
        messages.error(request, "Yêu cầu không hợp lệ hoặc không có file.")
        return redirect("add_student")


def import_csv_data(csv_file, request):
    data = csv.reader(csv_file.read().decode("utf-8").splitlines())

    try:
        headers = next(data)  # Bỏ qua dòng tiêu đề
    except StopIteration:
        raise ValueError("File CSV rỗng hoặc không hợp lệ")

    success_count = 0
    update_count = 0
    error_count = 0

    with transaction.atomic():
        for row in data:
            if len(row) != 8:  # Có 8 cột trong tệp CSV sau khi bỏ email
                raise ValueError("Số cột trong file CSV không đúng")

            ho, ten, password, gioi_tinh, dia_chi, ma_hoc_sinh, truong, lop = row

            # Kiểm tra nếu mã học sinh đã tồn tại
            if CustomUser.objects.filter(student_id=ma_hoc_sinh).exists():
                user = CustomUser.objects.get(student_id=ma_hoc_sinh)

                # Cập nhật thông tin cho tài khoản đã tồn tại
                user.first_name = ten
                user.last_name = ho
                user.gender = gioi_tinh
                user.address = dia_chi
                user.save()

                update_count += 1
                messages.info(
                    request, f"Cập nhật thông tin cho mã học sinh: {ma_hoc_sinh}"
                )
            else:
                # Nếu mã học sinh không tồn tại, thực hiện đăng ký mới
                response = account_register_auto(
                    ho, ten, password, gioi_tinh, dia_chi, ma_hoc_sinh, truong, lop
                )

                if isinstance(response, JsonResponse):
                    response_data = json.loads(response.content.decode("utf-8"))

                    if not response_data.get("success"):
                        error_message = response_data.get("errors", "Unknown error")
                        error_count += 1
                        messages.error(
                            request,
                            f"Không thể đăng ký học sinh với mã: {ma_hoc_sinh}. {error_message}",
                        )
                    else:
                        success_count += 1
                        print(f"Đăng ký thành công cho mã học sinh: {ma_hoc_sinh}")

    # Sau khi hoàn tất, gửi thông báo tổng kết
    messages.success(request, f"Đăng ký thành công {success_count} tài khoản mới.")
    if update_count > 0:
        messages.info(request, f"Đã cập nhật thông tin {update_count} tài khoản.")
    if error_count > 0:
        messages.error(request, f"Không thể đăng ký {error_count} tài khoản.")


def account_register_auto(
    ho, ten, password, gioi_tinh, dia_chi, ma_hoc_sinh, truong, lop
):
    try:
        # Kiểm tra xem mã học sinh đã tồn tại chưa trong bảng CustomUser
        if CustomUser.objects.filter(student_id=ma_hoc_sinh).exists():
            return JsonResponse(
                {"success": False, "errors": {"student_id": "Mã học sinh đã tồn tại"}}
            )

        # Tạo tài khoản mà không cần email
        user = CustomUser.objects.create_user(
            student_id=ma_hoc_sinh,
            password=password,
            user_type=3,  # Xác định đây là Student
            first_name=ten,
            last_name=ho,
        )
        user.gender = gioi_tinh
        user.address = dia_chi
        user.save()

        # Kiểm tra và gán trường học và lớp học
        try:
            course = Course.objects.get(name=truong)  # Lấy tên trường
        except Course.DoesNotExist:
            return JsonResponse(
                {"success": False, "errors": {"course": "Trường không tồn tại"}}
            )

        try:
            subject = Subject.objects.get(
                name=lop, course=course
            )  # Kiểm tra lớp học thuộc trường đó
        except Subject.DoesNotExist:
            return JsonResponse(
                {"success": False, "errors": {"subject": "Lớp không tồn tại"}}
            )

        # Kiểm tra xem học sinh đã tồn tại trong bảng Student chưa
        student, created = Student.objects.get_or_create(
            admin=user, defaults={"course": course, "session": subject}
        )

        if not created:
            # Nếu học sinh đã tồn tại, cập nhật thông tin course và session
            user.student.course = course
            user.student.session = subject
            user.save()

        # Tạo mã QR cho học sinh
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(ma_hoc_sinh)
        qr.make(fit=True)

        img = qr.make_image(fill="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        qr_image_name = f"{ma_hoc_sinh}_qr.png"

        # Lưu mã QR
        qr_dir = os.path.join(settings.MEDIA_ROOT, "QR")
        if not os.path.exists(qr_dir):
            os.makedirs(qr_dir)

        qr_path = os.path.join(qr_dir, qr_image_name)
        with open(qr_path, "wb") as f:
            f.write(buffer.getvalue())

        user.qr_code = os.path.join("QR", qr_image_name)
        user.save()

        return JsonResponse(
            {
                "success": True,
                "user_id": user.id,
                "message": "Đăng ký tài khoản thành công!",
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "errors": {"error": str(e)}})


from django.core.exceptions import ValidationError


def validate_csv_file(file):
    # Check if the file has a valid CSV extension
    ext = os.path.splitext(file.name)[-1].lower()
    if ext != ".csv":
        raise ValidationError("Định dạng file không hợp lệ. Chỉ chấp nhận file .csv.")

    # Check if the content type is correct (CSV file content type should be 'text/csv')
    if file.content_type != "text/csv":
        raise ValidationError(
            "Loại tệp không hợp lệ. Đảm bảo tệp là CSV với định dạng đúng."
        )


# thêm lịch dạy hàng loạt
def import_csv_teaching_schedule(request):
    if request.method == "POST" and "csv_file" in request.FILES:
        csv_file = request.FILES["csv_file"]

        try:
            validate_csv_file(csv_file)
        except ValidationError as e:
            messages.error(request, str(e))
            return redirect("add_session")

        try:
            data = csv.reader(csv_file.read().decode("utf-8").splitlines())
        except Exception:
            messages.error(request, "Lỗi khi đọc file CSV.")
            return redirect("add_session")

        try:
            headers = next(data)
        except StopIteration:
            messages.error(request, "File CSV rỗng hoặc không hợp lệ.")
            return redirect("add_session")

        success_count = 0
        error_count = 0
        today = timezone.now().date()

        # Tạo workbook và worksheet để lưu lịch dạy vào file Excel
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Lịch dạy"
        worksheet.append(
            ["Tên giáo viên", "Trường", "Lớp", "Ngày dạy", "Thời gian"]
        )  # Thêm tiêu đề cột

        with transaction.atomic():
            for row in data:
                if len(row) != 5:
                    messages.error(request, "Số cột trong file CSV không đúng.")
                    error_count += 1
                    continue

                email_giao_vien, truong, lop, ngay_day, thoi_gian = row

                try:
                    staff = Staff.objects.get(admin__email=email_giao_vien)
                except Staff.DoesNotExist:
                    messages.error(
                        request,
                        f"Không tìm thấy giáo viên có email: {email_giao_vien}.",
                    )
                    error_count += 1
                    continue

                try:
                    school = Course.objects.get(name=truong)
                except Course.DoesNotExist:
                    messages.error(request, f"Không tìm thấy trường: {truong}.")
                    error_count += 1
                    continue

                try:
                    subject = Subject.objects.get(name=lop, course=school)
                except Subject.DoesNotExist:
                    messages.error(
                        request, f"Không tìm thấy lớp: {lop} thuộc trường {truong}."
                    )
                    error_count += 1
                    continue

                try:
                    schedule_date = datetime.strptime(ngay_day, "%d/%m/%Y").date()
                    start_time = datetime.strptime(thoi_gian, "%Hh%M").time()
                except ValueError:
                    messages.error(
                        request,
                        f"Định dạng ngày hoặc giờ không hợp lệ: {ngay_day} {thoi_gian}.",
                    )
                    error_count += 1
                    continue

                if schedule_date >= today:
                    teaching_schedule = TeachingSchedule.objects.create(
                        staff=staff,
                        school_name=school,
                        class_name=subject,
                        schedule_date=schedule_date,
                        start_time=start_time,
                    )

                    # Thêm dữ liệu vào file Excel
                    worksheet.append(
                        [
                            teaching_schedule.staff.admin.last_name + " " + teaching_schedule.staff.admin.first_name,
                            truong,
                            lop,
                            schedule_date.strftime("%d/%m/%Y"),
                            start_time.strftime("%H:%M"),
                        ]
                    )
                    success_count += 1
                else:
                    messages.warning(
                        request,
                        f"Lịch dạy ngày {ngay_day} đã qua, không thêm vào hệ thống, chỉ thêm lịch dạy có ngày lớn hơn hoặc bằng với ngày hiện tại.",
                    )
                    error_count += 1

        # Lưu file Excel vào bộ nhớ
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)  # Đặt con trỏ về đầu file

        # Gửi email đính kèm file Excel cho giáo viên
        if success_count > 0:
            admin = staff.admin
            teacher_email = admin.email
            subject = "Thông báo danh sách lịch dạy mới"
            message = (
                f"Xin chào Thầy/Cô {admin.last_name} {admin.first_name},\n\n"
                f"Quý Thầy/Cô vừa được phân công danh sách lịch giảng dạy mới.\n"
                f"Quý Thầy/Cô kiểm tra lại toàn bộ lịch dạy trong file đính kèm và có những chuẩn bị cần thiết để buổi giảng dạy diễn ra thành công./.\n\n"
                "Trân trọng,\n \n"
                "HỆ THỐNG QUẢN LÝ ĐIỂM DANH BDU CÀ MAU"
            )

            email = EmailMessage(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [teacher_email],
            )

            # Đính kèm file Excel
            email.attach(
                "lich_day.xlsx",
                excel_file.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            email.send()

        messages.success(request, f"Đã thêm thành công {success_count} lịch dạy.")
        if error_count > 0:
            messages.error(
                request,
                f"Có {error_count} lịch dạy không được thêm do lỗi hoặc ngày giảng dạy đã qua.",
            )

        return redirect("manage_session")

    else:
        messages.error(request, "Không có file CSV nào được tải lên.")
        return redirect("add_session")


import logging

# xuất dữ liệu điểm danh ra file excel
logger = logging.getLogger(__name__)

def export_attendance(request):
    if request.method != "POST":
        return JsonResponse({"error": "Yêu cầu không hợp lệ"}, status=400)

    try:
        subject = request.POST.get("subject")
        session = request.POST.get("session")
        attendance_type = request.POST.get("attendance_type")

        # Date handling
        def convert_date(date_str):
            if "/" in date_str:
                return datetime.strptime(date_str, "%d/%m/%Y").strftime("%Y-%m-%d")
            return date_str

        def format_date(date_str):
            return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")

        attendance_records = None

        # Fetch attendance based on type (range or specific date)
        if attendance_type == "range":
            start_date = convert_date(request.POST.get("start_date"))
            end_date = convert_date(request.POST.get("end_date"))
            attendance_records = Attendance.objects.filter(
                subject_id=subject,
                session_id=session,
                date__range=(start_date, end_date),
            )
        else:
            attendance_date = convert_date(request.POST.get("attendance_date"))
            attendance_records = Attendance.objects.filter(
                subject_id=subject, session_id=session, date=attendance_date
            )

        if not attendance_records.exists():
            return JsonResponse(
                {
                    "error": "Không tìm thấy dữ liệu điểm danh cho lớp và khoảng thời gian đã chọn"
                },
                status=404,
            )

        # Present students
        present_students = AttendanceReport.objects.filter(
            attendance__in=attendance_records, status=True
        ).values_list("student", flat=True)

        # Absent students: Chỉ lấy học sinh thuộc lớp được chọn (session)
        students_absent = Student.objects.filter(session_id=session).exclude(
            id__in=present_students
        )

        # Data for present students with specific attendance dates
        data = [
            {
                "Tên học sinh": f"{report.student.admin.last_name} {report.student.admin.first_name}",
                "Mã học sinh": report.student.admin.student_id,
                "Lớp": Subject.objects.get(id=subject).name,
                "Ngày điểm danh": format_date(attendance.date.strftime("%Y-%m-%d")),
                "Trạng thái": "Có mặt" if report.status else "Chưa điểm danh",
            }
            for attendance in attendance_records
            for report in AttendanceReport.objects.filter(attendance=attendance)
        ]

        # Data for absent students, listing each date separately
        data1 = [
            {
                "Tên học sinh": f"{student.admin.last_name} {student.admin.first_name}",
                "Mã học sinh": student.admin.student_id,
                "Lớp": Subject.objects.get(id=student.session_id).name,
                "Ngày điểm danh": format_date(attendance.date.strftime("%Y-%m-%d")),
                "Trạng thái": "Chưa điểm danh",
            }
            for student in students_absent
            for attendance in attendance_records
        ]

        # Combine data and remove duplicates based on 'Mã học sinh' and 'Ngày điểm danh'
        combined_data = data + data1
        df = pd.DataFrame(combined_data).drop_duplicates(subset=["Mã học sinh", "Ngày điểm danh"])

        # Creating Excel file
        output = io.BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = "attachment; filename=attendance.xlsx"
        return response

    except ValueError as e:
        return JsonResponse({"error": "Định dạng ngày không hợp lệ"}, status=400)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


from django.core.paginator import Paginator


def manage_students(request):
    students_list = Student.objects.all()  # Lấy danh sách học sinh
    paginator = Paginator(students_list, 10)  # Hiển thị 10 học sinh mỗi trang
    page_number = request.GET.get("page")
    students = paginator.get_page(page_number)

    context = {
        "students": students,
        "page_title": "Quản lý học sinh",
    }
    return render(request, "hod_template/manage_student.html", context)
